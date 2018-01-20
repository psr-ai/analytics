import feedparser
import datetime
import time
import configparser
import logging
from openpyxl import Workbook, load_workbook
from DataAggregation.RSSFeeds.scrape_html.scrape import HTMLScraper


class RssFeedsCollector:

    def __init__(self, links_to_rss_feeds, html_scraper_configuration_object_path,
                 configuration_object_path=None, logs_path='logs.log', output_excel_path='output.xlsx',
                 wait_time_for_rerun=120, scrape_data_from_html_link=False):
        # Making logger
        log_filename = logs_path
        logging.basicConfig(filename=log_filename,level=logging.DEBUG, format='%(asctime)s %(message)s',
                            datefmt='%m/%d/%Y %I:%M:%S %p')

        # Making excel file and mentioning the headers of content
        self.output_excel_path = output_excel_path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.scrape_data_from_html_link = scrape_data_from_html_link

        if self.scrape_data_from_html_link:
            logging.info('Reading HMTL DOM Configuration object for HTML Scraping...')
            print('Creating HTML Scraping Object...')
            self.html_scraper = HTMLScraper(html_scraper_configuration_object_path)
            logging.info('DOM Configuration object successfully read.')
            print('DOM Configuration object successfully read.')
            self.ws.append(['Source', 'Title', 'Summary', 'PublishedOn', 'Link', 'Title_From_HTMLPage',
                            'Content_From_HTMLPage', 'Location_From_HTMLPage', 'Author_From_HTMLPage'])
        else:
            self.ws.append(['Source', 'Title', 'Summary', 'PublishedOn', 'Link'])

        self.wb.save(self.output_excel_path)

        # Objects to be used with each object of the class
        self.batch_of_data_to_append = []
        self.wait_time_for_rerun = wait_time_for_rerun
        self.rss_feeds_configuration = configparser.ConfigParser()
        if configuration_object_path is None:
            self.configuration_object_path = 'configuration.ini'
            # Making the configuration file
            logging.info('Creating the configuration file...')
            for link in links_to_rss_feeds:
                self.rss_feeds_configuration[link] = {
                    'last_entry_id': ''
                }
            self.update_configuration_file()
            logging.info('Configuration file created.')
        else:
            self.configuration_object_path = configuration_object_path

    def update_configuration_file(self):
        with open(self.configuration_object_path, 'w') as configfile:
                self.rss_feeds_configuration.write(configfile)

    def update_batch_of_data_to_append(self, entry, source, link):
        extracted_data = [source, entry['title'], entry['summary'], entry['published'], link]

        if self.scrape_data_from_html_link:
            if source in self.html_scraper.pages_structure:
                structured_data = self.html_scraper.get_page_structured_data(link, source)

                if 'heading' in structured_data:
                    extracted_data.append(structured_data['heading'])
                else:
                    extracted_data.append(None)

                if 'content' in structured_data:
                    extracted_data.append(structured_data['content'])
                else:
                    extracted_data.append(None)

                if 'location' in structured_data:
                    extracted_data.append(structured_data['location'])
                else:
                    extracted_data.append(None)

                if 'author' in structured_data:
                    extracted_data.append(structured_data['author'])
                else:
                    extracted_data.append(None)

        self.batch_of_data_to_append.append(extracted_data)

    def update_database(self):
        self.wb = load_workbook(self.output_excel_path)
        self.ws = self.wb.active
        for data in self.batch_of_data_to_append:
            self.ws.append(data)
        self.wb.save(self.output_excel_path)
        self.batch_of_data_to_append = []

    def run_rss_feed_collector(self):

        while True:
            logging.info("Running function at: " + str(datetime.datetime.now()))
            print("Running function at: " + str(datetime.datetime.now()))
            found_new_data = False
            self.rss_feeds_configuration.read(self.configuration_object_path)

            for source in self.rss_feeds_configuration.sections():
                feed_section = self.rss_feeds_configuration[source]
                id_of_first_link = ''
                feeds = feedparser.parse(source)
                for entry in feeds['entries']:
                    if entry['link'] != feed_section['last_entry_id'].replace('%%', '%'):
                        found_new_data = True
                        if id_of_first_link == '':
                            id_of_first_link = entry['link']
                        self.update_batch_of_data_to_append(entry, source, entry['link'])
                        logging.info('Found Post. Post Source: ' + source + ', Post Link: ' + entry['link'] +
                                     ', Post Title: ' + entry['title'])
                        print('Found Post. Post Source: ' + source + ', Post Link: ' + entry['link'] +
                              ', Post Title: ' + entry['title'])

                    else:
                        break

                if id_of_first_link != '':
                    feed_section['last_entry_id'] = id_of_first_link.replace('%', '%%')

            if found_new_data:
                logging.info('Updating database...')
                print('Updating database...')
                self.update_database()
                logging.info('Database updated.')
                print('Database updated.')
                logging.info('Updating configuration object...')
                print('Updating configuration object...')
                self.update_configuration_file()
                logging.info('Configuration object updated.')
                print('Configuration object updated.')

            logging.info('Sleeping for time: ' + str(self.wait_time_for_rerun) + ' seconds')
            print('Sleeping for time: ' + str(self.wait_time_for_rerun) + ' seconds')
            time.sleep(self.wait_time_for_rerun)


links_to_rss_feeds = ['http://feeds.reuters.com/Reuters/worldNews', 'http://feeds.reuters.com/reuters/INtopNews',
                      'http://feeds.reuters.com/reuters/INbusinessNews',
                      'http://feeds.reuters.com/reuters/INsouthAsiaNews', 'http://www.thehindu.com/news/?service=rss',
                      'http://www.bloomberg.com/politics/feeds/site.xml',
                      'http://www.ft.com/rss/companies/financial-services']
rss_collector = RssFeedsCollector(links_to_rss_feeds, 'news_websites_structure.ini', scrape_data_from_html_link=True)
rss_collector.run_rss_feed_collector()