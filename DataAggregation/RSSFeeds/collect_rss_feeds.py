import feedparser
import datetime
import time
import configparser
import logging
import os.path
from openpyxl import Workbook, load_workbook
from DataAggregation.Exceptions.RSSExceptions import CannotFindRSSFile
from DataAggregation.RSSFeeds.scrape_html.scrape import get_page_structured_data, pages_structure


class RssFeedsCollector:

    def __init__(self, configuration_object_path, wait_time_for_rerun=120,
                 excel_file_path='collect_rss_feeds/output.xlsx', scrape_data_from_html_link=True):

        if os.path.isfile(configuration_object_path):

            # Making excel file and mentioning the headers of content
            self.excel_file_path = excel_file_path
            self.wb = Workbook()
            self.ws = self.wb.active
            self.scrape_data_from_html_link = scrape_data_from_html_link

            if self.scrape_data_from_html_link:
                self.ws.append(['Source', 'Title', 'Summary', 'PublishedOn', 'Link', 'Title_From_HTMLPage',
                                'Content_From_HTMLPage', 'Location_From_HTMLPage', 'Author_From_HTMLPage'])
            else:
                self.ws.append(['Source', 'Title', 'Summary', 'PublishedOn', 'Link'])

            self.wb.save(self.excel_file_path)

            # Making logger
            log_filename = 'collect_rss_feeds/logs.log'
            logging.basicConfig(filename=log_filename,level=logging.DEBUG, format='%(asctime)s %(message)s',
                                datefmt='%m/%d/%Y %I:%M:%S %p')

            # Objects to be used with each object of the class
            self.batch_of_data_to_append = []
            self.rss_feeds_configuration = configparser.ConfigParser()
            self.configuration_object_path = configuration_object_path
            self.wait_time_for_rerun = wait_time_for_rerun
        else:
            raise CannotFindRSSFile(configuration_object_path)

    def update_batch_of_data_to_append(self, entry, source, link):
        extracted_data = [source, entry['title'], entry['summary'], entry['published'], link]

        if self.scrape_data_from_html_link:
            structured_data = {}
            if source in pages_structure:
                structured_data = get_page_structured_data(link, pages_structure[source])

            if 'heading' in structured_data:
                extracted_data.append(structured_data['heading'])
            else:
                extracted_data.append(None)

            if 'content' in structured_data:
                extracted_data.append(structured_data['content'])
            else:
                extracted_data.append(None)

            if 'location' in structured_data:
                extracted_data.append(structured_data['location'])
            else:
                extracted_data.append(None)

            if 'author' in structured_data:
                extracted_data.append(structured_data['author'])
            else:
                extracted_data.append(None)

        self.batch_of_data_to_append.append(extracted_data)

    def update_database(self):
        self.wb = load_workbook(self.excel_file_path)
        self.ws = self.wb.active
        for data in self.batch_of_data_to_append:
            self.ws.append(data)
        self.wb.save(self.excel_file_path)
        self.batch_of_data_to_append = []

    def run_rss_feed_collector(self):

        while True:
            logging.info("Running function at: " + str(datetime.datetime.now()))
            logging.error("Running function at: " + str(datetime.datetime.now()))
            print("Running function at: " + str(datetime.datetime.now()))
            found_new_data = False
            self.rss_feeds_configuration.read(self.configuration_object_path)

            for source in self.rss_feeds_configuration.sections():
                feed_section = self.rss_feeds_configuration[source]
                id_of_first_link = ''
                feeds = feedparser.parse(feed_section['link'])
                for entry in feeds['entries']:
                    if entry[feed_section['id_field']] != feed_section['last_entry_id']:
                        found_new_data = True
                        if id_of_first_link == '':
                            id_of_first_link = entry[feed_section['id_field']]
                        self.update_batch_of_data_to_append(entry, source, entry[feed_section['id_field']])
                        logging.info('Found Post. Post Source: ' + feed_section['link'] + ', Post Link: ' + entry[feed_section['id_field']] + ', Post Title: ' + entry['title'])
                        print('Found Post. Post Source: ' + feed_section['link'] + ', Post Link: ' + entry[feed_section['id_field']] + ', Post Title: ' + entry['title'])

                    else:
                        break

                if id_of_first_link != '':
                    feed_section['last_entry_id'] = id_of_first_link.replace('%', '%%')

            if found_new_data:
                logging.info('Updating database...')
                print('Updating database...')
                self.update_database()
                logging.info('Database updated.')
                print('Database updated.')
                logging.info('Updating configuration object...')
                print('Updating configuration object...')
                with open(self.configuration_object_path, 'w') as configfile:
                    self.rss_feeds_configuration.write(configfile)
                logging.info('Configuration object updated.')
                print('Configuration object updated.')

            logging.info('Sleeping for time: ' + str(self.wait_time_for_rerun) + ' seconds')
            print('Sleeping for time: ' + str(self.wait_time_for_rerun) + ' seconds')
            time.sleep(self.wait_time_for_rerun)

rss_collector = RssFeedsCollector('collect_rss_feeds/configuration.ini')
rss_collector.run_rss_feed_collector()